import argparse
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from processor import (
    escribir_reporte as escribir_reporte_caja,
    extraer_pdf as extraer_reporte_caja_pdf,
    normalizar_asignaciones_agencias,
)


ALIAS_COLUMNAS = {
    "CONTRATO": [
        "CONTRATO",
        "CONTRAT",
        "N CONTRATO",
        "NO CONTRATO",
        "NUMERO CONTRATO",
    ],
    "FECHA": ["FECHA", "FECHA VENTA", "FECHA EMISION"],
    "VENDEDOR": ["VENDEDOR", "ASESOR", "USUARIO VENDEDOR"],
    "CLIENTE": ["CLIENTE", "NOMBRE CLIENTE", "RAZON SOCIAL", "NOMBRES"],
    "MODELO": ["MODELO", "PRODUCTO", "DESCRIPCION", "ARTICULO"],
    "IMEI": ["IMEI", "IMEI 1", "SERIE", "SERIAL", "N SERIE", "NO SERIE"],
    "VENTAS": ["VENTAS", "VENTA", "VALOR VENTAS", "TOTAL VENTAS"],
    "ENTRADAS": ["ENTRADAS", "ENTRADA", "CUOTA INICIAL", "ENGANCHE"],
}

COLUMNAS_TV = [
    "CONTRATO",
    "FECHA",
    "VENDEDOR",
    "CLIENTE",
    "MODELO",
    "VENTAS",
    "ENTRADAS",
]

COLUMNAS_CELULAR = [
    "CONTRATO",
    "FECHA",
    "VENDEDOR",
    "CLIENTE",
    "MODELO",
    "IMEI",
    "VENTAS",
    "ENTRADAS",
]

PREFIJO_ARCHIVO_TEMPORAL = re.compile(r"^\d{10,}-[a-f0-9]{8}-", re.IGNORECASE)


def limpiar_celda(valor) -> str:
    if valor is None:
        return ""

    texto = re.sub(r"[\r\n]+", " ", str(valor))
    texto = re.sub(r"\s+", " ", texto).strip()
    return "" if texto.lower() in {"nan", "none", "null"} else texto


def normalizar_encabezado(valor) -> str:
    texto = unicodedata.normalize("NFD", limpiar_celda(valor).upper())
    texto = "".join(
        caracter for caracter in texto if unicodedata.category(caracter) != "Mn"
    )
    return re.sub(r"[^A-Z0-9]", "", texto)


def obtener_nombre_archivo(ruta: Path) -> str:
    return PREFIJO_ARCHIVO_TEMPORAL.sub("", ruta.name)


def calcular_hash_archivo(ruta: Path) -> str:
    digest = hashlib.sha256()
    with ruta.open("rb") as archivo:
        for bloque in iter(lambda: archivo.read(1024 * 1024), b""):
            digest.update(bloque)
    return digest.hexdigest()


def buscar_indice_columna(encabezados: list[str], alias: list[str]):
    alias_normalizados = [normalizar_encabezado(item) for item in alias]

    for indice, encabezado in enumerate(encabezados):
        if encabezado in alias_normalizados:
            return indice

    for indice, encabezado in enumerate(encabezados):
        if any(
            candidato and (candidato in encabezado or encabezado in candidato)
            for candidato in alias_normalizados
        ):
            return indice

    return None


def convertir_decimal(valor: str):
    texto = limpiar_celda(valor)
    if not texto:
        return None

    texto = re.sub(r"[^0-9,.-]", "", texto)
    if not texto or texto in {"-", ".", ","}:
        return None

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        partes = texto.split(",")
        texto = "".join(partes[:-1]) + "." + partes[-1]

    try:
        return round(float(texto), 2)
    except ValueError:
        return None


def listar_pdfs(directorio: str | None) -> list[Path]:
    if not directorio:
        return []

    ruta = Path(directorio)
    if not ruta.exists():
        return []

    return sorted(
        archivo
        for archivo in ruta.iterdir()
        if archivo.is_file() and archivo.suffix.lower() == ".pdf"
    )


def crear_error(tipo: str, archivo: str, pagina, fila, motivo: str, detalle=""):
    return {
        "TIPO": tipo,
        "ARCHIVO": archivo,
        "PAGINA": pagina,
        "FILA": fila,
        "MOTIVO": motivo,
        "DETALLE": limpiar_celda(detalle),
    }


def extraer_ventas_pdf(ruta_pdf: Path, tipo: str):
    registros = []
    errores = []
    archivo = obtener_nombre_archivo(ruta_pdf)
    archivo_hash = calcular_hash_archivo(ruta_pdf)
    columnas = COLUMNAS_CELULAR if tipo == "CELULAR" else COLUMNAS_TV
    tablas_encontradas = 0

    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            for numero_pagina, pagina in enumerate(pdf.pages, start=1):
                for tabla in pagina.extract_tables() or []:
                    if not tabla or len(tabla) < 2:
                        continue

                    encabezados = [normalizar_encabezado(item) for item in tabla[0]]
                    indices = {
                        columna: buscar_indice_columna(
                            encabezados,
                            ALIAS_COLUMNAS[columna],
                        )
                        for columna in columnas
                    }

                    columnas_faltantes = [
                        columna for columna, indice in indices.items() if indice is None
                    ]
                    if columnas_faltantes:
                        continue

                    tablas_encontradas += 1

                    for numero_fila, fila in enumerate(tabla[1:], start=2):
                        valores = [limpiar_celda(valor) for valor in (fila or [])]
                        texto_fila = " ".join(valores).upper()

                        if not texto_fila.strip() or "TOTALES GENERALES" in texto_fila:
                            continue

                        registro = {}
                        for columna, indice in indices.items():
                            registro[columna] = (
                                valores[indice] if indice is not None and indice < len(valores) else ""
                            )

                        columnas_identificacion = [
                            columna
                            for columna in columnas
                            if columna not in {"VENTAS", "ENTRADAS"}
                        ]
                        if all(
                            not registro.get(columna)
                            for columna in columnas_identificacion
                        ):
                            # Los reportes cierran cada tabla con una fila que solo
                            # contiene los totales monetarios.
                            continue

                        faltantes = [
                            columna
                            for columna in columnas
                            if not registro.get(columna)
                        ]
                        if faltantes:
                            errores.append(
                                crear_error(
                                    tipo,
                                    archivo,
                                    numero_pagina,
                                    numero_fila,
                                    "DATOS_OBLIGATORIOS_FALTANTES",
                                    ", ".join(faltantes),
                                )
                            )
                            continue

                        ventas = convertir_decimal(registro["VENTAS"])
                        entradas = convertir_decimal(registro["ENTRADAS"])
                        if ventas is None or entradas is None:
                            errores.append(
                                crear_error(
                                    tipo,
                                    archivo,
                                    numero_pagina,
                                    numero_fila,
                                    "VALORES_MONETARIOS_INVALIDOS",
                                    f"VENTAS={registro['VENTAS']}; ENTRADAS={registro['ENTRADAS']}",
                                )
                            )
                            continue

                        registro["VENTAS"] = ventas
                        registro["ENTRADAS"] = entradas
                        registro["ARCHIVO"] = archivo
                        registro["ARCHIVO_HASH"] = archivo_hash
                        registros.append(registro)
    except Exception as exc:
        errores.append(
            crear_error(tipo, archivo, None, None, "PDF_NO_LEIDO", str(exc))
        )
        return registros, errores

    if tablas_encontradas == 0:
        errores.append(
            crear_error(
                tipo,
                archivo,
                None,
                None,
                "FORMATO_NO_RECONOCIDO",
                f"No se encontro una tabla con las columnas de ventas de {tipo.lower()}.",
            )
        )

    return registros, errores


def extraer_ventas(pdfs: list[Path], tipo: str):
    registros = []
    errores = []

    for pdf in pdfs:
        registros_pdf, errores_pdf = extraer_ventas_pdf(pdf, tipo)
        registros.extend(registros_pdf)
        errores.extend(errores_pdf)

    return registros, errores


def aplicar_estilo_tabla(ws, columnas_texto=None):
    columnas_texto = columnas_texto or []
    azul = "1F4E78"
    blanco = "FFFFFF"
    verde = "E2F0D9"
    borde = Border(
        left=Side(style="thin", color="B7C3D0"),
        right=Side(style="thin", color="B7C3D0"),
        top=Side(style="thin", color="B7C3D0"),
        bottom=Side(style="thin", color="B7C3D0"),
    )

    for celda in ws[1]:
        celda.font = Font(bold=True, color=blanco)
        celda.fill = PatternFill("solid", fgColor=azul)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila in ws.iter_rows():
        for celda in fila:
            if celda.value is not None:
                celda.border = borde
                celda.alignment = Alignment(vertical="center", wrap_text=True)

    if ws.max_row > 1:
        for celda in ws[ws.max_row]:
            celda.font = Font(bold=True)
            celda.fill = PatternFill("solid", fgColor=verde)

    for indice in columnas_texto:
        for fila in range(2, ws.max_row + 1):
            ws.cell(fila, indice).number_format = "@"

    for columna in range(1, ws.max_column + 1):
        letra = get_column_letter(columna)
        maximo = max(
            (len(str(celda.value)) for celda in ws[letra] if celda.value is not None),
            default=10,
        )
        ws.column_dimensions[letra].width = min(max(maximo + 2, 12), 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def escribir_hoja_ventas(wb, nombre: str, registros: list[dict], columnas: list[str]):
    if nombre in wb.sheetnames:
        del wb[nombre]

    ws = wb.create_sheet(nombre)
    ws.append(columnas)

    for registro in registros:
        ws.append([registro.get(columna, "") for columna in columnas])

    fila_total = [None] * len(columnas)
    indice_ventas = columnas.index("VENTAS")
    indice_entradas = columnas.index("ENTRADAS")
    fila_total[max(0, indice_ventas - 1)] = "TOTAL"
    fila_total[indice_ventas] = round(
        sum(registro["VENTAS"] for registro in registros), 2
    )
    fila_total[indice_entradas] = round(
        sum(registro["ENTRADAS"] for registro in registros), 2
    )
    ws.append(fila_total)

    for indice in (indice_ventas + 1, indice_entradas + 1):
        for fila in range(2, ws.max_row + 1):
            ws.cell(fila, indice).number_format = '$#,##0.00'

    columnas_texto = [columnas.index("CONTRATO") + 1]
    if "IMEI" in columnas:
        columnas_texto.append(columnas.index("IMEI") + 1)
    aplicar_estilo_tabla(ws, columnas_texto)


def procesar_reportes_caja_con_detalle(
    reportes_caja: list[Path],
    salida: Path,
    asignaciones_agencias=None,
):
    registros = []
    no_leidas_por_archivo = defaultdict(list)
    detalle_archivos = []
    asignaciones = normalizar_asignaciones_agencias(asignaciones_agencias or [])

    for pdf in reportes_caja:
        extraidos, no_leidas = extraer_reporte_caja_pdf(pdf, asignaciones)
        nombre_archivo = obtener_nombre_archivo(pdf)
        archivo_hash = calcular_hash_archivo(pdf)
        for registro in extraidos:
            registro["ARCHIVO"] = nombre_archivo
            registro["ARCHIVO_HASH"] = archivo_hash
        registros.extend(extraidos)

        if no_leidas:
            no_leidas_por_archivo[pdf.name].extend(no_leidas)

        detalle_archivos.append(
            {
                "archivo": nombre_archivo,
                "archivoHash": archivo_hash,
                "registros": len(extraidos),
                "noLeidas": len(no_leidas),
            }
        )

    if not registros:
        raise ValueError(
            "No se extrajo ningun registro valido de los PDFs de caja seleccionados."
        )

    escribir_reporte_caja(registros, salida, no_leidas_por_archivo)

    return (
        {
            "registros": len(registros),
            "noLeidas": sum(len(items) for items in no_leidas_por_archivo.values()),
            "archivos": detalle_archivos,
        },
        registros,
    )


def procesar(
    reportes_caja: list[Path],
    ventas_tv_pdfs: list[Path],
    ventas_celular_pdfs: list[Path],
    salida: Path,
    asignaciones_agencias=None,
    salida_datos: Path | None = None,
):
    if not reportes_caja and not ventas_tv_pdfs and not ventas_celular_pdfs:
        raise ValueError(
            "Selecciona al menos un PDF de caja, ventas TV o ventas celular."
        )

    if reportes_caja:
        resumen_caja, registros_caja = procesar_reportes_caja_con_detalle(
            reportes_caja,
            salida,
            asignaciones_agencias or [],
        )
        wb = load_workbook(salida)
        hoja_inicial = None
    else:
        resumen_caja = {"registros": 0, "noLeidas": 0, "archivos": []}
        registros_caja = []
        wb = Workbook()
        hoja_inicial = wb.active

    ventas_tv, errores_tv = extraer_ventas(ventas_tv_pdfs, "TV")
    ventas_celular, errores_celular = extraer_ventas(
        ventas_celular_pdfs,
        "CELULAR",
    )

    if ventas_tv_pdfs and not ventas_tv:
        raise ValueError(
            "No se extrajo ninguna venta de TV. Revisa que los PDFs correspondan al formato TV."
        )

    if ventas_celular_pdfs and not ventas_celular:
        raise ValueError(
            "No se extrajo ninguna venta de celular. Revisa que los PDFs incluyan la columna IMEI."
        )

    for nombre_hoja in ("Filas no leidas", "Errores ventas"):
        if nombre_hoja in wb.sheetnames:
            del wb[nombre_hoja]

    escribir_hoja_ventas(wb, "Ventas TV", ventas_tv, COLUMNAS_TV)
    escribir_hoja_ventas(
        wb,
        "Ventas Celular",
        ventas_celular,
        COLUMNAS_CELULAR,
    )
    if hoja_inicial is not None and hoja_inicial.title in wb.sheetnames:
        del wb[hoja_inicial.title]
    errores_ventas = [*errores_tv, *errores_celular]
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)

    if salida_datos:
        salida_datos.parent.mkdir(parents=True, exist_ok=True)
        with salida_datos.open("w", encoding="utf-8") as archivo_datos:
            json.dump(
                {
                    "registrosCaja": registros_caja,
                    "ventasTv": ventas_tv,
                    "ventasCelular": ventas_celular,
                },
                archivo_datos,
                ensure_ascii=False,
            )

    return {
        "registros": resumen_caja.get("registros", 0),
        "registrosCaja": resumen_caja.get("registros", 0),
        "ventasTv": len(ventas_tv),
        "ventasCelular": len(ventas_celular),
        "noLeidasCaja": resumen_caja.get("noLeidas", 0),
        "erroresVentas": len(errores_ventas),
        "noLeidas": resumen_caja.get("noLeidas", 0) + len(errores_ventas),
        "archivos": resumen_caja.get("archivos", []),
    }


def main():
    parser = argparse.ArgumentParser(
        description="Genera el cierre de caja con anexos de ventas TV y celular."
    )
    parser.add_argument("--output", required=True)
    parser.add_argument("--reportes-caja-dir", required=True)
    parser.add_argument("--ventas-tv-dir")
    parser.add_argument("--ventas-celular-dir")
    parser.add_argument("--data-output")
    parser.add_argument("--asignaciones-agencias", default="[]")
    args = parser.parse_args()

    try:
        asignaciones_agencias = json.loads(args.asignaciones_agencias)
    except json.JSONDecodeError as exc:
        raise ValueError(
            "Las asignaciones de agencias no tienen un formato valido."
        ) from exc

    resultado = procesar(
        listar_pdfs(args.reportes_caja_dir),
        listar_pdfs(args.ventas_tv_dir),
        listar_pdfs(args.ventas_celular_dir),
        Path(args.output),
        asignaciones_agencias,
        Path(args.data_output) if args.data_output else None,
    )
    print(json.dumps(resultado, ensure_ascii=False))


if __name__ == "__main__":
    main()
