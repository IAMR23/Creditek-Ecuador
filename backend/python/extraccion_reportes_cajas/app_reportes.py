import re
import sys
import threading
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import pdfplumber
import customtkinter as ctk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# CONFIGURACIÓN DE EXTRACCIÓN
# ==========================================================

MAPEO_AGENCIAS = {
    "ALEXFER": "NUEVA AURORA",
    "DAMIZA": "CAUPICHO",
    "CHAVICTK": "SANGOLQUI",
}

PRODUCTOS_VALIDOS = ["CREDITV", "UPHONE"]

PATRON_INICIO_FILA = re.compile(
    r"^\s*(?P<contrato>\d+)\s+"
    r"(?P<fecha>\d{1,2}/\d{1,2}/\d{2})\s+"
    r"(?P<hora>\d{1,2}:\d{1,2})(?:\s+(?P<ampm>AM|PM))?\s+"
    r"(?P<vendedor>\S+)\s+"
    r"(?P<usuario>\S+)\s+"
    r"(?P<resto>.+?)\s*$",
    re.IGNORECASE,
)

PATRON_POSIBLE_REGISTRO = re.compile(
    r"^\d+\s+\d{1,2}/\d{1,2}/\d{2}\s+"
)

PATRON_TOKEN_PAGO = re.compile(r"^\d+(?:[.,]\d{1,3})?$")

PATRON_SOLO_CUOTAS = re.compile(
    r"^[\d,\s]+(?:DE(?:\s+\d+)?)?$",
    re.IGNORECASE,
)


# ==========================================================
# FUNCIONES DE PROCESAMIENTO
# ==========================================================

def limpiar_linea(linea: str) -> str:
    return re.sub(r"\s+", " ", str(linea)).strip()


def normalizar_producto(texto: str, nombre_archivo: str) -> str:
    base = f"{nombre_archivo} {texto}".upper()

    if "UPHONE" in base or "UPH" in base:
        return "UPHONE"

    if "CREDITV" in base or "CREDI-TV" in base or "CREDI TV" in base:
        return "CREDITV"

    return "OTRO"


def obtener_agencia(usuario_cobrador: str) -> str:
    usuario = (usuario_cobrador or "").upper().strip()

    for clave, agencia in MAPEO_AGENCIAS.items():
        if clave in usuario:
            return agencia

    return "OTROS"


def convertir_pago(valor: str) -> float:
    valor = valor.strip()

    if "," in valor and "." not in valor:
        valor = valor.replace(",", ".")
    else:
        valor = valor.replace(",", "")

    return round(float(valor), 2)


def extraer_texto_pdf(ruta_pdf: Path) -> str:
    paginas_texto = []

    with pdfplumber.open(ruta_pdf) as pdf:
        for page in pdf.pages:
            texto = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            paginas_texto.append(texto)

    return "\n".join(paginas_texto)


def separar_cliente_pago_cuotas(resto: str):
    tokens = resto.split()
    indice_pago = None

    for i, token in enumerate(tokens):
        token_limpio = token.strip()

        if PATRON_TOKEN_PAGO.match(token_limpio):
            indice_pago = i
            break

    if indice_pago is None:
        return None

    cliente = " ".join(tokens[:indice_pago]).strip()
    pago = tokens[indice_pago].strip()
    n_cuotas = " ".join(tokens[indice_pago + 1:]).strip()

    if not cliente:
        return None

    return cliente, pago, n_cuotas


def unir_lineas_cortadas(lineas: list[str]) -> list[str]:
    resultado = []

    for linea in lineas:
        linea = limpiar_linea(linea)

        if not linea:
            continue

        es_registro = bool(PATRON_POSIBLE_REGISTRO.match(linea))

        if es_registro:
            resultado.append(linea)
            continue

        if resultado and PATRON_SOLO_CUOTAS.match(linea):
            resultado[-1] = limpiar_linea(resultado[-1] + " " + linea)
            continue

        resultado.append(linea)

    return resultado


def extraer_pdf(ruta_pdf: Path) -> tuple[list[dict], list[str]]:
    registros = []
    no_leidas = []

    texto_completo = extraer_texto_pdf(ruta_pdf)
    producto = normalizar_producto(texto_completo, ruta_pdf.name)

    lineas_originales = texto_completo.splitlines()
    lineas = unir_lineas_cortadas(lineas_originales)

    for linea_limpia in lineas:
        m_inicio = PATRON_INICIO_FILA.match(linea_limpia)

        if not m_inicio:
            if PATRON_POSIBLE_REGISTRO.match(linea_limpia):
                no_leidas.append(linea_limpia)
            continue

        resto = m_inicio.group("resto").strip()
        separado = separar_cliente_pago_cuotas(resto)

        if not separado:
            no_leidas.append(linea_limpia)
            continue

        cliente, pago_txt, n_cuotas = separado

        contrato = int(m_inicio.group("contrato"))
        fecha = m_inicio.group("fecha")
        hora = m_inicio.group("hora")
        ampm = m_inicio.group("ampm") or ""
        fecha_hora = f"{fecha} {hora} {ampm}".strip()

        vendedor = m_inicio.group("vendedor").upper().strip()
        usuario = m_inicio.group("usuario").upper().strip()

        try:
            pagos = convertir_pago(pago_txt)
        except ValueError:
            no_leidas.append(linea_limpia)
            continue

        registros.append({
            "CONTRATO": contrato,
            "FECHA": fecha_hora,
            "VENDEDOR": vendedor,
            "USUARIO COBRADOR": usuario,
            "CLIENTE": cliente.upper().strip(),
            "PAGOS CUOTAS": pagos,
            "Nº CUOTAS": n_cuotas.upper().strip(),
            "PRODUCTO": producto,
            "AGENCIA": obtener_agencia(usuario),
            "ARCHIVO": ruta_pdf.name,
        })

    return registros, no_leidas


# ==========================================================
# GENERACIÓN DE EXCEL
# ==========================================================

def aplicar_estilos(ws):
    azul = "1F4E78"
    verde = "70AD47"
    rojo = "C00000"
    blanco = "FFFFFF"

    borde = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if cell.value is not None:
                cell.border = borde

    for row in ws.iter_rows():
        values = [
            str(c.value).upper() if c.value is not None else ""
            for c in row
        ]

        if "CONTRATO" in values or ("AGENCIA" in values and "TOTAL" in values):
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True, color=blanco)
                    cell.fill = PatternFill("solid", fgColor=azul)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        elif row[0].value in PRODUCTOS_VALIDOS or row[0].value == "RESUMEN GENERAL":
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True, color=blanco, size=12)
                    cell.fill = PatternFill("solid", fgColor=verde)

        elif row[0].value == "TOTAL":
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True, color=blanco)
                    cell.fill = PatternFill("solid", fgColor=rojo)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row, col).value
            if isinstance(valor, (int, float)):
                ws.cell(row, col).number_format = '#,##0.00'

    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)

        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A4"


def aplicar_estilos_no_leidas(ws):
    azul = "1F4E78"
    blanco = "FFFFFF"

    borde = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = Font(bold=True, color=blanco)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"


def escribir_reporte(
    registros: list[dict],
    salida: Path,
    no_leidas_por_archivo: dict[str, list[str]]
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(["RESUMEN GENERAL"])
    ws.append(["AGENCIA", "UPHONE", "CREDITV", "TOTAL"])

    agencias = ["NUEVA AURORA", "CAUPICHO", "SANGOLQUI", "OTROS"]
    resumen = defaultdict(lambda: defaultdict(float))

    for r in registros:
        producto = r["PRODUCTO"]
        agencia = r["AGENCIA"]

        if producto in PRODUCTOS_VALIDOS:
            resumen[agencia][producto] += r["PAGOS CUOTAS"]

    for agencia in agencias:
        uphone = round(resumen[agencia]["UPHONE"], 2)
        creditv = round(resumen[agencia]["CREDITV"], 2)
        total = round(uphone + creditv, 2)
        ws.append([agencia, uphone, creditv, total])

    total_uphone = round(sum(resumen[a]["UPHONE"] for a in resumen), 2)
    total_creditv = round(sum(resumen[a]["CREDITV"] for a in resumen), 2)
    total_general = round(total_uphone + total_creditv, 2)

    ws.append(["TOTAL", total_uphone, total_creditv, total_general])
    ws.append([])

    columnas = [
        "CONTRATO",
        "FECHA",
        "VENDEDOR",
        "USUARIO COBRADOR",
        "CLIENTE",
        "PAGOS CUOTAS",
        "Nº CUOTAS",
        "AGENCIA",
        "ARCHIVO",
    ]

    for producto in PRODUCTOS_VALIDOS:
        datos_producto = [
            r for r in registros
            if r["PRODUCTO"] == producto
        ]

        ws.append([producto])
        ws.append(columnas)

        for r in datos_producto:
            ws.append([r[c] for c in columnas])

        total_producto = round(
            sum(r["PAGOS CUOTAS"] for r in datos_producto),
            2
        )

        ws.append([None, None, None, None, "TOTAL", total_producto])
        ws.append([])
        ws.append([])

    aplicar_estilos(ws)

    ws_no = wb.create_sheet("Filas no leídas")
    ws_no.append(["ARCHIVO", "LÍNEA NO LEÍDA"])

    for archivo, lineas in no_leidas_por_archivo.items():
        for linea in lineas:
            ws_no.append([archivo, linea])

    aplicar_estilos_no_leidas(ws_no)

    wb.save(salida)


# ==========================================================
# APP DE ESCRITORIO
# ==========================================================

class AppReportes(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Generador de Reportes PDF a Excel")
        self.geometry("980x680")
        self.minsize(900, 620)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.pdfs_seleccionados: list[Path] = []
        self.ruta_salida: Path | None = None

        self.crear_interfaz()

    def crear_interfaz(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        # Header
        header = ctk.CTkFrame(self, corner_radius=0, fg_color="#111827")
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        titulo = ctk.CTkLabel(
            header,
            text="Generador de Reportes",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="#FFFFFF"
        )
        titulo.grid(row=0, column=0, padx=28, pady=(22, 4), sticky="w")

        subtitulo = ctk.CTkLabel(
            header,
            text="Selecciona PDFs, define una ruta de salida y genera el Excel consolidado automáticamente.",
            font=ctk.CTkFont(size=14),
            text_color="#CBD5E1"
        )
        subtitulo.grid(row=1, column=0, padx=28, pady=(0, 22), sticky="w")

        # Contenedor principal
        main = ctk.CTkFrame(self, corner_radius=18)
        main.grid(row=1, column=0, padx=24, pady=20, sticky="ew")
        main.grid_columnconfigure(1, weight=1)

        # Selección de PDFs
        lbl_pdfs = ctk.CTkLabel(
            main,
            text="PDFs seleccionados",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        lbl_pdfs.grid(row=0, column=0, padx=20, pady=(20, 8), sticky="w")

        self.entry_pdfs = ctk.CTkEntry(
            main,
            placeholder_text="Ningún PDF seleccionado",
            height=38
        )
        self.entry_pdfs.grid(row=0, column=1, padx=12, pady=(20, 8), sticky="ew")

        btn_pdfs = ctk.CTkButton(
            main,
            text="Seleccionar PDFs",
            height=38,
            command=self.seleccionar_pdfs
        )
        btn_pdfs.grid(row=0, column=2, padx=20, pady=(20, 8))

        # Ruta salida
        lbl_salida = ctk.CTkLabel(
            main,
            text="Archivo Excel salida",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        lbl_salida.grid(row=1, column=0, padx=20, pady=8, sticky="w")

        self.entry_salida = ctk.CTkEntry(
            main,
            placeholder_text="Selecciona dónde guardar el Excel",
            height=38
        )
        self.entry_salida.grid(row=1, column=1, padx=12, pady=8, sticky="ew")

        btn_salida = ctk.CTkButton(
            main,
            text="Guardar como",
            height=38,
            command=self.seleccionar_salida
        )
        btn_salida.grid(row=1, column=2, padx=20, pady=8)

        # Resumen
        stats = ctk.CTkFrame(main, corner_radius=14, fg_color="#1F2937")
        stats.grid(row=2, column=0, columnspan=3, padx=20, pady=(18, 20), sticky="ew")
        stats.grid_columnconfigure((0, 1, 2), weight=1)

        self.lbl_total_pdfs = ctk.CTkLabel(
            stats,
            text="PDFs: 0",
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.lbl_total_pdfs.grid(row=0, column=0, padx=20, pady=16)

        self.lbl_registros = ctk.CTkLabel(
            stats,
            text="Registros: 0",
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.lbl_registros.grid(row=0, column=1, padx=20, pady=16)

        self.lbl_no_leidas = ctk.CTkLabel(
            stats,
            text="No leídas: 0",
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.lbl_no_leidas.grid(row=0, column=2, padx=20, pady=16)

        # Acciones
        acciones = ctk.CTkFrame(self, corner_radius=18)
        acciones.grid(row=2, column=0, padx=24, pady=(0, 16), sticky="ew")
        acciones.grid_columnconfigure(0, weight=1)

        self.progress = ctk.CTkProgressBar(acciones)
        self.progress.grid(row=0, column=0, padx=20, pady=(20, 8), sticky="ew")
        self.progress.set(0)

        self.lbl_estado = ctk.CTkLabel(
            acciones,
            text="Listo para procesar.",
            font=ctk.CTkFont(size=13),
            text_color="#CBD5E1"
        )
        self.lbl_estado.grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")

        botones = ctk.CTkFrame(acciones, fg_color="transparent")
        botones.grid(row=0, column=1, rowspan=2, padx=20, pady=20)

        self.btn_generar = ctk.CTkButton(
            botones,
            text="Generar Excel",
            height=44,
            width=180,
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.iniciar_generacion
        )
        self.btn_generar.grid(row=0, column=0, padx=6)

        # Logs
        logs_frame = ctk.CTkFrame(self, corner_radius=18)
        logs_frame.grid(row=3, column=0, padx=24, pady=(0, 24), sticky="nsew")
        logs_frame.grid_columnconfigure(0, weight=1)
        logs_frame.grid_rowconfigure(1, weight=1)

        lbl_logs = ctk.CTkLabel(
            logs_frame,
            text="Registro del proceso",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        lbl_logs.grid(row=0, column=0, padx=20, pady=(18, 8), sticky="w")

        self.txt_logs = ctk.CTkTextbox(logs_frame, height=260)
        self.txt_logs.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.log("Aplicación iniciada correctamente.")

    def log(self, mensaje: str):
        hora = datetime.now().strftime("%H:%M:%S")
        self.txt_logs.insert("end", f"[{hora}] {mensaje}\n")
        self.txt_logs.see("end")

    def seleccionar_pdfs(self):
        archivos = filedialog.askopenfilenames(
            title="Selecciona los PDFs",
            filetypes=[
                ("Archivos PDF", "*.pdf"),
                ("Todos los archivos", "*.*")
            ]
        )

        if not archivos:
            return

        self.pdfs_seleccionados = [Path(a) for a in archivos]
        self.entry_pdfs.delete(0, "end")
        self.entry_pdfs.insert(0, f"{len(self.pdfs_seleccionados)} PDF(s) seleccionado(s)")
        self.lbl_total_pdfs.configure(text=f"PDFs: {len(self.pdfs_seleccionados)}")

        self.log(f"Se seleccionaron {len(self.pdfs_seleccionados)} PDF(s).")

        if not self.ruta_salida:
            carpeta = self.pdfs_seleccionados[0].parent
            nombre = f"REPORTE_GENERADO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            self.ruta_salida = carpeta / nombre
            self.entry_salida.delete(0, "end")
            self.entry_salida.insert(0, str(self.ruta_salida))

    def seleccionar_salida(self):
        archivo = filedialog.asksaveasfilename(
            title="Guardar reporte como",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel", "*.xlsx")
            ],
            initialfile=f"REPORTE_GENERADO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )

        if not archivo:
            return

        self.ruta_salida = Path(archivo)
        self.entry_salida.delete(0, "end")
        self.entry_salida.insert(0, str(self.ruta_salida))
        self.log(f"Ruta de salida seleccionada: {self.ruta_salida}")

    def iniciar_generacion(self):
        if not self.pdfs_seleccionados:
            messagebox.showwarning("Faltan PDFs", "Selecciona al menos un PDF.")
            return

        if not self.ruta_salida:
            messagebox.showwarning("Falta salida", "Selecciona la ruta de salida del Excel.")
            return

        self.btn_generar.configure(state="disabled")
        self.progress.set(0)
        self.lbl_estado.configure(text="Procesando PDFs...")
        self.lbl_registros.configure(text="Registros: 0")
        self.lbl_no_leidas.configure(text="No leídas: 0")

        hilo = threading.Thread(target=self.generar_excel, daemon=True)
        hilo.start()

    def generar_excel(self):
        registros = []
        todas_no_leidas = defaultdict(list)

        total_pdfs = len(self.pdfs_seleccionados)

        try:
            for i, pdf in enumerate(self.pdfs_seleccionados, start=1):
                self.after(0, self.log, f"Procesando: {pdf.name}")

                extraidos, no_leidas = extraer_pdf(pdf)

                registros.extend(extraidos)

                if no_leidas:
                    todas_no_leidas[pdf.name].extend(no_leidas)

                progreso = i / total_pdfs

                self.after(0, self.progress.set, progreso)
                self.after(0, self.lbl_estado.configure, {
                    "text": f"Procesando {i}/{total_pdfs}: {pdf.name}"
                })
                self.after(0, self.lbl_registros.configure, {
                    "text": f"Registros: {len(registros)}"
                })
                self.after(0, self.lbl_no_leidas.configure, {
                    "text": f"No leídas: {sum(len(v) for v in todas_no_leidas.values())}"
                })

                self.after(
                    0,
                    self.log,
                    f"{pdf.name}: {len(extraidos)} registros extraídos, {len(no_leidas)} no leídas."
                )

            if not registros:
                raise ValueError("No se extrajo ningún registro válido de los PDFs seleccionados.")

            escribir_reporte(registros, self.ruta_salida, todas_no_leidas)

            total_no_leidas = sum(len(v) for v in todas_no_leidas.values())

            self.after(0, self.progress.set, 1)
            self.after(0, self.lbl_estado.configure, {
                "text": "Excel generado correctamente."
            })
            self.after(0, self.log, f"Excel generado: {self.ruta_salida}")
            self.after(0, self.log, f"Total registros: {len(registros)}")
            self.after(0, self.log, f"Total filas no leídas: {total_no_leidas}")

            self.after(
                0,
                messagebox.showinfo,
                "Proceso completado",
                f"Excel generado correctamente.\n\n"
                f"Ruta:\n{self.ruta_salida}\n\n"
                f"Registros: {len(registros)}\n"
                f"Filas no leídas: {total_no_leidas}"
            )

        except Exception as e:
            self.after(0, self.progress.set, 0)
            self.after(0, self.lbl_estado.configure, {
                "text": "Error al generar el Excel."
            })
            self.after(0, self.log, f"ERROR: {str(e)}")
            self.after(0, messagebox.showerror, "Error", str(e))

        finally:
            self.after(0, self.btn_generar.configure, {
                "state": "normal"
            })


if __name__ == "__main__":
    app = AppReportes()
    app.mainloop()