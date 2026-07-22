import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


MODULO_DIR = Path(__file__).resolve().parent
BACKEND_DIR = MODULO_DIR.parents[1]
PDFS_DIR = BACKEND_DIR / "pdfs"
sys.path.insert(0, str(MODULO_DIR))

import ventas_cierre_processor as processor


class VentasCierreProcessorTest(unittest.TestCase):
    def test_extrae_los_dos_formatos_de_ventas(self):
        tv, errores_tv = processor.extraer_ventas(
            [PDFS_DIR / "ReportUphone_CierreCajaVentas_IAMRS2.pdf"],
            "TV",
        )
        celulares, errores_celulares = processor.extraer_ventas(
            [PDFS_DIR / "AURORA 17 AL 19.pdf"],
            "CELULAR",
        )

        self.assertEqual(len(tv), 2)
        self.assertEqual(errores_tv, [])
        self.assertGreater(len(celulares), 0)
        self.assertEqual(errores_celulares, [])
        self.assertNotIn("IMEI", tv[0])
        self.assertTrue(celulares[0]["IMEI"])
        self.assertRegex(tv[0]["ARCHIVO_HASH"], r"^[a-f0-9]{64}$")
        self.assertRegex(celulares[0]["ARCHIVO_HASH"], r"^[a-f0-9]{64}$")

    def test_adjunta_hojas_de_ventas_al_excel_de_caja(self):
        def crear_excel_base(_reportes, salida, _asignaciones):
            wb = Workbook()
            wb.active.title = "Reporte"
            wb.create_sheet("Filas no leidas")
            wb.save(salida)
            return (
                {"registros": 3, "noLeidas": 1, "archivos": []},
                [
                    {
                        "CONTRATO": 1,
                        "FECHA": "1/1/26 10:00 AM",
                        "VENDEDOR": "VENDEDOR",
                        "USUARIO COBRADOR": "COBRADOR",
                        "CLIENTE": "CLIENTE",
                        "PAGOS CUOTAS": 10,
                        "Nro CUOTAS": "1",
                        "PRODUCTO": "UPHONE",
                        "AGENCIA": "CAUPICHO",
                        "ARCHIVO": "caja.pdf",
                    }
                ],
            )

        with tempfile.TemporaryDirectory() as temp_dir:
            salida = Path(temp_dir) / "cierre.xlsx"
            salida_datos = Path(temp_dir) / "control-financiero.json"

            with patch.object(
                processor,
                "procesar_reportes_caja_con_detalle",
                side_effect=crear_excel_base,
            ):
                resumen = processor.procesar(
                    [PDFS_DIR / "AURORA 17 AL 19.pdf"],
                    [PDFS_DIR / "ReportUphone_CierreCajaVentas_IAMRS2.pdf"],
                    [PDFS_DIR / "AURORA 17 AL 19.pdf"],
                    salida,
                    [],
                    salida_datos,
                )

            wb = load_workbook(salida, data_only=False)
            self.assertEqual(
                wb.sheetnames,
                [
                    "Reporte",
                    "Ventas TV",
                    "Ventas Celular",
                ],
            )
            self.assertEqual(
                [celda.value for celda in wb["Ventas TV"][1]],
                processor.COLUMNAS_TV,
            )
            self.assertEqual(
                [celda.value for celda in wb["Ventas Celular"][1]],
                processor.COLUMNAS_CELULAR,
            )
            self.assertEqual(resumen["registrosCaja"], 3)
            self.assertEqual(resumen["ventasTv"], 2)
            self.assertGreater(resumen["ventasCelular"], 0)
            with salida_datos.open("r", encoding="utf-8") as archivo_datos:
                datos = json.load(archivo_datos)
            self.assertEqual(len(datos["registrosCaja"]), 1)
            self.assertEqual(len(datos["ventasTv"]), 2)


if __name__ == "__main__":
    unittest.main()
