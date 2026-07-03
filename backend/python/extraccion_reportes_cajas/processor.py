import argparse
import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MAPEO_AGENCIAS = {
    "ALEXFER": "NUEVA AURORA",
    "DAMIZA": "CAUPICHO",
    "CHAVICTK": "SANGOLQUI",
}

PRODUCTOS_VALIDOS = ["CREDITV", "UPHONE"]

PATRON_INICIO_FILA = re.compile(
    r"^\s*(?P<contrato>\d+)\s+"
    r"(?P<fecha>\d{1,2}/\d{1,2}/\d{2})\s+"
    r"(?P<hora>\d{1,2}:\d{1,2})(?:\s+(?P<ampm>AM|PM))?\s+"
    r"(?P<vendedor>\S+)\s+"
    r"(?P<usuario>\S+)\s+"
    r"(?P<resto>.+?)\s*$",
    re.IGNORECASE,
)

PATRON_POSIBLE_REGISTRO = re.compile(r"^\d+\s+\d{1,2}/\d{1,2}/\d{2}\s+")
PATRON_TOKEN_PAGO = re.compile(r"^\d+(?:[.,]\d{1,3})?$")
PATRON_SOLO_CUOTAS = re.compile(r"^[\d,\s]+(?:DE(?:\s+\d+)?)?$", re.IGNORECASE)


def limpiar_linea(linea: str) -> str:
    return re.sub(r"\s+", " ", str(linea)).strip()


def normalizar_producto(texto: str, nombre_archivo: str) -> str:
    base = f"{nombre_archivo} {texto}".upper()

    if "UPHONE" in base or "UPH" in base:
        return "UPHONE"

    if "CREDITV" in base or "CREDI-TV" in base or "CREDI TV" in base:
        return "CREDITV"

    return "OTRO"


def construir_fecha_iso(dia: str, mes: str, anio: str) -> str:
    anio_num = int(anio)
    if anio_num < 100:
        anio_num += 2000

    try:
        return date(anio_num, int(mes), int(dia)).isoformat()
    except ValueError:
        return ""


def normalizar_fecha_asignacion(fecha: str) -> str:
    fecha = (fecha or "").strip()

    if re.match(r"^\d{4}-\d{2}-\d{2}$", fecha):
        try:
            return date.fromisoformat(fecha).isoformat()
        except ValueError:
            return ""

    match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})$", fecha)
    if not match:
        return ""

    dia, mes, anio = match.groups()
    return construir_fecha_iso(dia, mes, anio)


def obtener_fechas_posibles(fecha: str) -> list[str]:
    fecha = (fecha or "").strip()

    if re.match(r"^\d{4}-\d{2}-\d{2}$", fecha):
        normalizada = normalizar_fecha_asignacion(fecha)
        return [normalizada] if normalizada else []

    match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})$", fecha)
    if not match:
        return []

    primero, segundo, anio = match.groups()
    fechas = []

    for dia, mes in ((primero, segundo), (segundo, primero)):
        normalizada = construir_fecha_iso(dia, mes, anio)
        if normalizada and normalizada not in fechas:
            fechas.append(normalizada)

    return fechas


def normalizar_asignaciones_agencias(raw) -> dict[str, dict[str, str]]:
    asignaciones = defaultdict(dict)

    if not raw:
        return asignaciones

    for item in raw:
        fecha = normalizar_fecha_asignacion(str(item.get("fecha", "")))
        usuario = str(item.get("usuario", "")).upper().strip()
        agencia = str(item.get("agencia", "")).upper().strip()

        if fecha and usuario and agencia:
            asignaciones[fecha][usuario] = agencia

    return asignaciones


def obtener_agencia(
    usuario_cobrador: str,
    fecha: str = "",
    asignaciones_agencias: dict[str, dict[str, str]] | None = None,
) -> str:
    usuario = (usuario_cobrador or "").upper().strip()
    fechas_posibles = obtener_fechas_posibles(fecha)

    if asignaciones_agencias:
        for fecha_key in fechas_posibles:
            for clave, agencia in asignaciones_agencias.get(fecha_key, {}).items():
                if clave in usuario:
                    return agencia

    for clave, agencia in MAPEO_AGENCIAS.items():
        if clave in usuario:
            return agencia

    return "OTROS"


def convertir_pago(valor: str) -> float:
    valor = valor.strip()

    if "," in valor and "." not in valor:
        valor = valor.replace(",", ".")
    else:
        valor = valor.replace(",", "")

    return round(float(valor), 2)


def extraer_texto_pdf(ruta_pdf: Path) -> str:
    paginas_texto = []

    with pdfplumber.open(ruta_pdf) as pdf:
        for page in pdf.pages:
            texto = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            paginas_texto.append(texto)

    return "\n".join(paginas_texto)


def separar_cliente_pago_cuotas(resto: str):
    tokens = resto.split()
    indice_pago = None

    for i, token in enumerate(tokens):
        token_limpio = token.strip()

        if PATRON_TOKEN_PAGO.match(token_limpio):
            indice_pago = i
            break

    if indice_pago is None:
        return None

    cliente = " ".join(tokens[:indice_pago]).strip()
    pago = tokens[indice_pago].strip()
    n_cuotas = " ".join(tokens[indice_pago + 1 :]).strip()

    if not cliente:
        return None

    return cliente, pago, n_cuotas


def unir_lineas_cortadas(lineas: list[str]) -> list[str]:
    resultado = []

    for linea in lineas:
        linea = limpiar_linea(linea)

        if not linea:
            continue

        es_registro = bool(PATRON_POSIBLE_REGISTRO.match(linea))

        if es_registro:
            resultado.append(linea)
            continue

        if resultado and PATRON_SOLO_CUOTAS.match(linea):
            resultado[-1] = limpiar_linea(resultado[-1] + " " + linea)
            continue

        resultado.append(linea)

    return resultado


def extraer_pdf(
    ruta_pdf: Path,
    asignaciones_agencias: dict[str, dict[str, str]] | None = None,
) -> tuple[list[dict], list[str]]:
    registros = []
    no_leidas = []

    texto_completo = extraer_texto_pdf(ruta_pdf)
    producto = normalizar_producto(texto_completo, ruta_pdf.name)

    lineas = unir_lineas_cortadas(texto_completo.splitlines())

    for linea_limpia in lineas:
        m_inicio = PATRON_INICIO_FILA.match(linea_limpia)

        if not m_inicio:
            if PATRON_POSIBLE_REGISTRO.match(linea_limpia):
                no_leidas.append(linea_limpia)
            continue

        separado = separar_cliente_pago_cuotas(m_inicio.group("resto").strip())

        if not separado:
            no_leidas.append(linea_limpia)
            continue

        cliente, pago_txt, n_cuotas = separado
        fecha = m_inicio.group("fecha")
        hora = m_inicio.group("hora")
        ampm = m_inicio.group("ampm") or ""

        try:
            pagos = convertir_pago(pago_txt)
        except ValueError:
            no_leidas.append(linea_limpia)
            continue

        usuario = m_inicio.group("usuario").upper().strip()

        registros.append(
            {
                "CONTRATO": int(m_inicio.group("contrato")),
                "FECHA": f"{fecha} {hora} {ampm}".strip(),
                "VENDEDOR": m_inicio.group("vendedor").upper().strip(),
                "USUARIO COBRADOR": usuario,
                "CLIENTE": cliente.upper().strip(),
                "PAGOS CUOTAS": pagos,
                "Nro CUOTAS": n_cuotas.upper().strip(),
                "PRODUCTO": producto,
                "AGENCIA": obtener_agencia(usuario, fecha, asignaciones_agencias),
                "ARCHIVO": ruta_pdf.name,
            }
        )

    return registros, no_leidas


def aplicar_estilos(ws):
    azul = "1F4E78"
    verde = "70AD47"
    rojo = "C00000"
    blanco = "FFFFFF"

    borde = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if cell.value is not None:
                cell.border = borde

    for row in ws.iter_rows():
        values = [str(c.value).upper() if c.value is not None else "" for c in row]

        if "CONTRATO" in values or ("AGENCIA" in values and "TOTAL" in values):
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True, color=blanco)
                    cell.fill = PatternFill("solid", fgColor=azul)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        elif row[0].value in PRODUCTOS_VALIDOS or row[0].value == "RESUMEN GENERAL":
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True, color=blanco, size=12)
                    cell.fill = PatternFill("solid", fgColor=verde)

        elif row[0].value == "TOTAL":
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True, color=blanco)
                    cell.fill = PatternFill("solid", fgColor=rojo)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row, col).value
            if isinstance(valor, (int, float)):
                ws.cell(row, col).number_format = "#,##0.00"

    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)

        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A4"


def aplicar_estilos_no_leidas(ws):
    azul = "1F4E78"
    blanco = "FFFFFF"
    borde = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = Font(bold=True, color=blanco)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"


def escribir_reporte(registros: list[dict], salida: Path, no_leidas_por_archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(["RESUMEN GENERAL"])
    ws.append(["AGENCIA", "UPHONE", "CREDITV", "TOTAL"])

    agencias = ["NUEVA AURORA", "CAUPICHO", "SANGOLQUI", "OTROS"]
    resumen = defaultdict(lambda: defaultdict(float))

    for r in registros:
        if r["PRODUCTO"] in PRODUCTOS_VALIDOS:
            resumen[r["AGENCIA"]][r["PRODUCTO"]] += r["PAGOS CUOTAS"]

    for agencia in agencias:
        uphone = round(resumen[agencia]["UPHONE"], 2)
        creditv = round(resumen[agencia]["CREDITV"], 2)
        ws.append([agencia, uphone, creditv, round(uphone + creditv, 2)])

    total_uphone = round(sum(resumen[a]["UPHONE"] for a in resumen), 2)
    total_creditv = round(sum(resumen[a]["CREDITV"] for a in resumen), 2)
    ws.append(["TOTAL", total_uphone, total_creditv, round(total_uphone + total_creditv, 2)])
    ws.append([])

    columnas = [
        "CONTRATO",
        "FECHA",
        "VENDEDOR",
        "USUARIO COBRADOR",
        "CLIENTE",
        "PAGOS CUOTAS",
        "Nro CUOTAS",
        "AGENCIA",
        "ARCHIVO",
    ]

    for producto in PRODUCTOS_VALIDOS:
        datos_producto = [r for r in registros if r["PRODUCTO"] == producto]
        ws.append([producto])
        ws.append(columnas)

        for r in datos_producto:
            ws.append([r[c] for c in columnas])

        total_producto = round(sum(r["PAGOS CUOTAS"] for r in datos_producto), 2)
        ws.append([None, None, None, None, "TOTAL", total_producto])
        ws.append([])
        ws.append([])

    aplicar_estilos(ws)

    ws_no = wb.create_sheet("Filas no leidas")
    ws_no.append(["ARCHIVO", "LINEA NO LEIDA"])

    for archivo, lineas in no_leidas_por_archivo.items():
        for linea in lineas:
            ws_no.append([archivo, linea])

    aplicar_estilos_no_leidas(ws_no)
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)


def procesar(pdfs: list[Path], salida: Path, asignaciones_agencias=None):
    registros = []
    todas_no_leidas = defaultdict(list)
    detalle_archivos = []
    asignaciones = normalizar_asignaciones_agencias(asignaciones_agencias or [])

    for pdf in pdfs:
        extraidos, no_leidas = extraer_pdf(pdf, asignaciones)
        registros.extend(extraidos)

        if no_leidas:
            todas_no_leidas[pdf.name].extend(no_leidas)

        detalle_archivos.append(
            {
                "archivo": pdf.name,
                "registros": len(extraidos),
                "noLeidas": len(no_leidas),
            }
        )

    if not registros:
        raise ValueError("No se extrajo ningun registro valido de los PDFs seleccionados.")

    escribir_reporte(registros, salida, todas_no_leidas)

    return {
        "registros": len(registros),
        "noLeidas": sum(len(v) for v in todas_no_leidas.values()),
        "archivos": detalle_archivos,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--asignaciones-agencias", default="[]")
    parser.add_argument("pdfs", nargs="+")
    args = parser.parse_args()

    try:
        asignaciones_agencias = json.loads(args.asignaciones_agencias)
    except json.JSONDecodeError as exc:
        raise ValueError("Las asignaciones de agencias no tienen un formato valido.") from exc

    pdfs = [Path(pdf) for pdf in args.pdfs]
    salida = Path(args.output)
    resultado = procesar(pdfs, salida, asignaciones_agencias)
    print(json.dumps(resultado, ensure_ascii=False))


if __name__ == "__main__":
    main()
